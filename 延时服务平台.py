import os
import traceback

import numpy as np
from openpyxl import load_workbook

import pandas as pd
from openpyxl.utils import get_column_letter
from pandas import DataFrame, ExcelWriter
import xlwt

import RequestsUtil

cookie = "Hm_lvt_27d444d6f19a5ce4b7d2bbca8959f7dc=1681803236,1683637379,1683637567; Hm_lpvt_27d444d6f19a5ce4b7d2bbca8959f7dc=1683637576"
authorization = "Bearer xihmUi5kh9Dn9gKsrhYhjVHeHVa73N3gIsULg4o2zGx%2BWREh1fk4wi2oT5v7C4rMacsnWlkCI5yV39TrjK0gyy1lAyqtKRFlb5xHb4gSyoY5r5BA4BRnZS%2FYblMBJMbx"


def getHeader(method, path):
    return {
        "authority": "sctfxq.yanshifuwu.com",
        "method": method,
        "path": path,
        "scheme": "https",
        "accept": "application/json, text/plain, */*",
        "accept-encoding": "gzip, deflate, br",
        "accept-language": "zh-CN,zh;q=0.9",
        "authorization": authorization,
        "cookie": cookie,
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "Windows",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "content-type": "multipart/form-data; boundary=----WebKitFormBoundary2ERI3BgRbWbz26uE",
        "sec-fetch-site": "same-origin",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36",
    }



def getKc(id):
    re = RequestsUtil.requestGet("https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/selectionResult?id=273&grade=&min_enroll=&max_enroll=&what_week_num=&course_name=&teacher_name=",
                            getHeader("GET", "/Teaching/TeachingStatistics/selectionResult?id="+str(id)+"&grade=&min_enroll=&max_enroll=&what_week_num=&course_name=&teacher_name="),
                            {'id': id})
    print(re)
    reJson = re.json()
    print(reJson)
    return reJson['data']

def downMd(class_name, class_id):
    RequestsUtil.downFile("F://课程//" + class_name + ".xls", "https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/studentCourseByClassEx",
                          getHeader("get", "/Teaching/TeachingStatistics/studentCourseByClassEx?timetable_course_class_id="+str(class_id)+"&student_name=&page=1&limit=1000"),
                            {"timetable_course_class_id": class_id, "page": 1, "limit": 1000}
                          )

# 基于文件名和表格名删除一些行和一些列，注意没有备份。
# flsh是指文件名flname和表格名sheetname
def del_flsh_rows_cols(flname, sheetname, rowd, cold):  # 基于文件名和表格名删除一些行和一些列
    """基于文件名和表格名删除一些行和一些列
    要删的行序数放在rowd表格中，要删的列序数放在cold表格中
    本程序的关键是删除的行或列序数都必须是从大的开始删除，这样才不会乱序"""

    sheet = pd.read_excel(flname, sheet_name=sheetname)
    pd.set_option('display.max_columns', 10000, 'display.max_rows', 10000)
    #课程名称	年级班级	学生姓名	学号/ID	家长姓名	家长手机号	物料费
    #del sheet['课程名称', '年级班级', '家长姓名', '家长手机号', '物料费']
    #s = sheet.drop(labels=['学号/ID', '家长姓名', '物料费'], axis=1)
    #课程名称	年级班级	学生姓名	家长手机号
    nj = sheet['年级班级'].astype('str').values
    xm = sheet['学生姓名'].astype('str').values
    arr = []
    index = 0
    for item in sheet['家长手机号'].astype('str').values:
        arr.append(ccjz[str(nj[index] + xm[index]).replace('\n', '')])
        index = index + 1

    df1=pd.DataFrame({'课程名称':sheet['课程名称'].astype('str'),
                      '年级班级':sheet['年级班级'].astype('str'),
                      '学生姓名':sheet['学生姓名'].astype('str'),
                      '家长手机号':arr,
                      })
    pd.set_option('display.max_colwidth',2000) #最大列字符数
    df1=df1.set_index('课程名称')#以id列作为索引列，即为第一列
    df1.to_excel(flname)#将数据框转为Excel并保存
    #da = DataFrame(s).apply(lambda s: s.apply('{0}'.format))
    #da.astype("str")
    #da.to_excel(flname, sheet_name=sheetname, index=False, header=True)

ccjz = {}

def readPath(path, path1):
    paths = os.listdir(path)
    for p in paths:
        with pd.ExcelWriter(path1 + p + "x") as writer:
            to_excel_auto_column_weight(getDp(path + p, 'Sheet1'), writer, f'Sheet1')

def getDp(path, sheetname):
    sheet = pd.read_excel(path, sheet_name=sheetname)
    df1=pd.DataFrame({'课程名称':sheet['课程名称'].astype('str'),
                      '年级班级':sheet['年级班级'].astype('str'),
                      '学生姓名':sheet['学生姓名'].astype('str'),
                      '家长手机号':sheet['家长手机号'].astype('str'),
                      })
    return df1





def getTalHome():
    sheet = pd.read_excel("F:\\全校名单.xls", sheet_name="Worksheet")
    nj = sheet['年级班级'].astype('str').values
    name = sheet['学生姓名'].astype('str').values
    jz = sheet['家长电话'].astype('str').values
    arr = {}
    index = 0
    for item in jz:
        arr[str(nj[index] + name[index]).replace("\n", "")]=item.replace('nan', '').replace('.0', '').replace('\n', '')
        index = index + 1
    return arr

def uploadFileKc(path):
    kcs = getKc()
    paths = os.listdir(path)
    for p in paths:
        for kc in kcs:
            if(kc['class_name'] + ".xls" == p):
                #print(kc['class_name'])
                re = RequestsUtil.uploadFile(path + p, "https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/studentResultByClassIm",
                                        getHeader("post", "https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/studentResultByClassIm"),
                                        {"timetable_course_class_id": str(kc['id'])})
                print(re.json())
        #RequestsUtil.uploadFile(path + p, "")


def test():
    df1 = pd.DataFrame({'课程名称':['课程名称','课程名称'],
                  '家长手机号':['111111111111','11111111111111111111111'],
                  })
    pd.set_option('display.max_colwidth',2000) #最大列字符数
    df1=df1.set_index('课程名称')#以id列作为索引列，即为第一列
    df1.to_excel("F:\\a.xls")#将数据框转为Excel并保存

def to_excel_auto_column_weight(df: pd.DataFrame, writer: ExcelWriter, sheet_name):
    """DataFrame保存为excel并自动设置列宽"""
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    #  计算表头的字符宽度
    column_widths = (
        df.columns.to_series().apply(lambda x: len(x.encode('gbk'))).values
    )
    #  计算每列的最大字符宽度
    max_widths = (
        df.astype(str).applymap(lambda x: len(x.encode('gbk'))).agg(max).values
    )
    # 计算整体最大宽度
    widths = np.max([column_widths, max_widths], axis=0)
    # 设置列宽
    worksheet = writer.sheets[sheet_name]
    for i, width in enumerate(widths, 1):
        # openpyxl引擎设置字符宽度时会缩水0.5左右个字符，所以干脆+2使左右都空出一个字宽。
        worksheet.column_dimensions[get_column_letter(i)].width = width + 2


# 获取班级
def getOrg():
    re = RequestsUtil.requestGet("https://sctfxq.yanshifuwu.com/Internal/Teacher/Search/getOrgNew",
                            getHeader("GET", "/Internal/Teacher/Search/getOrgNew"),
                                 {})
    cs = re.json()['data'][0]['children'][0]['children']
    classs = []
    for c in cs:
        for c1 in c['children']:
            classs.append(c1)
    print(classs)
    return classs

# 下载课程表
def downClass(id):


    myClass = getOrg()
    for cl in myClass:
        try:
            print(cl)
            re = RequestsUtil.requestGet("https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/studentResult",
                                         getHeader("GET", "/Teaching/TeachingStatistics/studentResult"),
                                         {'id': id,
                                          'org_id': cl['id'],
                                          'status': 3,
                                          'pay_status': '',
                                          'student_name': '',
                                          'page': 1,
                                          'limit': 200})
            #班级
            col1 = []
            col2 = []
            col3 = []
            col4 = []
            col5 = []
            for item in re.json()['data']['data']:
                for item1 in item['course']:
                    col1.append(str(item1['class_name']))
                    col2.append(str(item['student_name']))
                    col3.append(str(item['parent_mobile']))
                    col4.append(str(item1['build_name']))
                    col5.append(str(item1['time_cn']).replace("<br>", "  "))
            df1=pd.DataFrame({'班级': col1,
                              '学生姓名': col2,
                              '家长电话': col3,
                              '场地': col4,
                              '上课时间': col5,
                              })
            with pd.ExcelWriter("F://班级//" + cl['name'] + ".xlsx") as writer:
                to_excel_auto_column_weight(df1, writer, f'sheet')
        except BaseException:
            print(cl)
            traceback.print_exc()

def getKcNow(id):
    kcs = getKc(273)
    for kc in kcs:
        print(kc['class_name'])
        print(kc['id'])
        #https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/studentCourseByClass?timetable_course_class_id=12469&student_name=&page=1&limit=40

        try:
            print(kc)
            re = RequestsUtil.requestGet("https://sctfxq.yanshifuwu.com/Teaching/TeachingStatistics/studentCourseByClass",
                                         getHeader("GET", "/Teaching/TeachingStatistics/studentCourseByClass"),
                                         {'timetable_course_class_id': kc['id'],
                                          'page': 1,
                                          'limit': 1000})
            #班级
            col1 = []
            col2 = []
            col3 = []
            col4 = []
            if(re.json()['data']['data'].__len__() == 0) :
                continue
            class_name = str(kc['class_name'])
            if(class_name == "1班" or class_name == "2班") :
                class_name = str(kc['course_name']) + class_name
            for item in re.json()['data']['data']:
                col1.append(class_name)
                org_names = str(item['org_name']).split("/")
                col2.append(org_names[org_names.__len__() - 1])
                col3.append(str(item['student_name']))
                col4.append(str(item['parent_mobile']))
            df1=pd.DataFrame({'课程名称': col1,
                              '年级班级': col2,
                              '学生姓名': col3,
                              '家长手机号': col4,
                              })
            with pd.ExcelWriter("F://课程//" + class_name + ".xlsx") as writer:
                to_excel_auto_column_weight(df1, writer, f'sheet')
        except BaseException:
            print(kc)
            traceback.print_exc()


if __name__ == "__main__":
    #uploadFileKc("F://课程1//")
    #print("周1".startswith("周"))
    #ccjz = getTalHome()

    #downClass(273)

    #uploadFileKc("F://课程//")
    #downMd("街舞一班", 842)
    kcs = getKcNow(273)
    # for kc in kcs:
    #    print(kc['class_name'])
    #    print(kc['id'])
    #    #if(kc['time_cn'].startswith("周一")):
    #    downMd(kc['class_name'], kc['id'])
    #downMd("街舞基础1班", 842)
    #filePath = os.path("F://课程")
    #print(os.listdir("F://课程"))
    #del_flsh_rows_cols("F://课程//乒乓球1班.xls", "Worksheet", [], [1, 2, 5, 6, 7])

